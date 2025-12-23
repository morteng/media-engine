"""
Excel Builder

Generates XLSX spreadsheets from data definitions.
Applies theme styling automatically.

Features:
- Styled headers and data rows
- Currency and percentage formatting
- Charts (bar, line)
- Formulas
- Conditional formatting
- Multiple worksheets
"""

from dataclasses import dataclass, field
from pathlib import Path
from typing import TYPE_CHECKING, Any, Dict, List

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from ..brand import BrandContext
    from ..core.theme import Theme

# Import centralized color utility
from ..brand.colors import hex_to_argb


@dataclass
class Column:
    """Column definition."""

    name: str
    width: float = 12
    format: str = "general"  # general, currency, percent, date, number
    formula: str = ""  # Optional formula template


@dataclass
class SheetData:
    """Data for a worksheet."""

    name: str
    title: str = ""
    columns: List[Column] = field(default_factory=list)
    rows: List[List[Any]] = field(default_factory=list)
    include_chart: bool = False
    chart_type: str = "bar"  # bar, line
    chart_title: str = ""
    chart_data_range: str = ""  # e.g., "A1:B10"


class XLSXBuilder:
    """Builds Excel spreadsheets with theme styling."""

    def __init__(self, theme: "Theme" = None, brand: "BrandContext" = None):
        """
        Initialize XLSX builder.

        Args:
            theme: Legacy Theme for styling (deprecated, use brand instead)
            brand: BrandContext for unified brand access (recommended)
        """
        self.brand = brand

        if brand:
            # Use BrandContext
            self.color_primary = brand.hex_to_argb(brand.get_brand_color("primary"))
            self.color_secondary = brand.hex_to_argb(brand.get_brand_color("secondary"))
            self.color_accent = brand.hex_to_argb(brand.get_brand_color("accent"))
            self.color_bg = brand.hex_to_argb(brand.get_color("background.primary") or "#ffffff")
            self.color_muted = brand.hex_to_argb(brand.get_color("text.muted") or "#666666")

            self.font_heading = brand.get_system_font("heading")
            self.font_body = brand.get_system_font("body")
        else:
            from ..core.theme import COPPER_AND_CREAM

            self.theme = theme or COPPER_AND_CREAM

            self.color_primary = hex_to_argb(self.theme.colors.primary)
            self.color_secondary = hex_to_argb(self.theme.colors.secondary)
            self.color_accent = hex_to_argb(self.theme.colors.accent)
            self.color_bg = hex_to_argb(self.theme.colors.background)
            self.color_muted = hex_to_argb(self.theme.colors.muted)

            # Default to Helvetica Neue for legacy mode
            self.font_heading = "Helvetica Neue"
            self.font_body = "Helvetica Neue"

        # Create workbook
        self.wb = Workbook()
        self._setup_styles()

        # Remove default sheet (we'll add our own)
        self._default_removed = False

    def _setup_styles(self):
        """Setup named styles for consistent formatting."""
        # Header style
        header_style = NamedStyle(name="header")
        header_style.font = Font(name=self.font_heading, size=11, bold=True, color="FFFFFF")
        header_style.fill = PatternFill(
            start_color=self.color_primary, end_color=self.color_primary, fill_type="solid"
        )
        header_style.alignment = Alignment(horizontal="center", vertical="center")
        header_style.border = Border(bottom=Side(style="thin", color=self.color_primary))
        self.wb.add_named_style(header_style)

        # Title style
        title_style = NamedStyle(name="title")
        title_style.font = Font(
            name=self.font_heading, size=16, bold=True, color=self.color_primary
        )
        title_style.alignment = Alignment(horizontal="left", vertical="center")
        self.wb.add_named_style(title_style)

        # Data style
        data_style = NamedStyle(name="data")
        data_style.font = Font(name=self.font_body, size=10, color=self.color_primary)
        data_style.alignment = Alignment(horizontal="left", vertical="center")
        self.wb.add_named_style(data_style)

        # Number style
        number_style = NamedStyle(name="number")
        number_style.font = Font(name=self.font_body, size=10, color=self.color_primary)
        number_style.alignment = Alignment(horizontal="right")
        number_style.number_format = "#,##0"
        self.wb.add_named_style(number_style)

        # Currency style
        currency_style = NamedStyle(name="currency")
        currency_style.font = Font(name=self.font_body, size=10, color=self.color_primary)
        currency_style.alignment = Alignment(horizontal="right")
        currency_style.number_format = '#,##0.00 "USD"'
        self.wb.add_named_style(currency_style)

        # Percent style
        percent_style = NamedStyle(name="percent")
        percent_style.font = Font(name=self.font_body, size=10, color=self.color_primary)
        percent_style.alignment = Alignment(horizontal="right")
        percent_style.number_format = "0.0%"
        self.wb.add_named_style(percent_style)

        # Input style (editable cells)
        input_style = NamedStyle(name="input")
        input_style.font = Font(name=self.font_body, size=10, color=self.color_primary)
        input_style.fill = PatternFill(start_color="FFFEF9", end_color="FFFEF9", fill_type="solid")
        input_style.border = Border(
            left=Side(style="thin", color=self.color_accent),
            right=Side(style="thin", color=self.color_accent),
            top=Side(style="thin", color=self.color_accent),
            bottom=Side(style="thin", color=self.color_accent),
        )
        input_style.alignment = Alignment(horizontal="right")
        self.wb.add_named_style(input_style)

        # Highlight style (key results)
        highlight_style = NamedStyle(name="highlight")
        highlight_style.font = Font(name=self.font_heading, size=11, bold=True, color="FFFFFF")
        highlight_style.fill = PatternFill(
            start_color=self.color_accent, end_color=self.color_accent, fill_type="solid"
        )
        highlight_style.alignment = Alignment(horizontal="center", vertical="center")
        self.wb.add_named_style(highlight_style)

        # Alternating row style
        alt_row_style = NamedStyle(name="alt_row")
        alt_row_style.font = Font(name=self.font_body, size=10, color=self.color_primary)
        alt_row_style.fill = PatternFill(
            start_color="FFF7F4F1", end_color="FFF7F4F1", fill_type="solid"
        )
        self.wb.add_named_style(alt_row_style)

    def _remove_default_sheet(self):
        """Remove the default empty sheet."""
        if not self._default_removed and "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]
            self._default_removed = True

    def add_sheet(
        self,
        name: str,
        title: str = "",
        columns: List[Column] = None,
        rows: List[List[Any]] = None,
    ):
        """
        Add a worksheet with data.

        Args:
            name: Sheet name
            title: Optional title above data
            columns: Column definitions
            rows: Data rows
        """
        self._remove_default_sheet()

        ws = self.wb.create_sheet(title=name)
        current_row = 1

        # Add title if provided
        if title:
            ws.cell(row=current_row, column=1, value=title)
            ws.cell(row=current_row, column=1).style = "title"
            ws.merge_cells(
                start_row=current_row,
                start_column=1,
                end_row=current_row,
                end_column=len(columns) if columns else 1,
            )
            current_row += 2  # Leave a blank row

        # Add headers
        if columns:
            for col_idx, col in enumerate(columns, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=col.name)
                cell.style = "header"
                ws.column_dimensions[get_column_letter(col_idx)].width = col.width

            current_row += 1

        # Add data rows
        if rows:
            for row_idx, row_data in enumerate(rows):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=current_row, column=col_idx, value=value)

                    # Apply format based on column definition
                    if columns and col_idx <= len(columns):
                        col_format = columns[col_idx - 1].format
                        if col_format == "currency":
                            cell.style = "currency"
                        elif col_format == "percent":
                            cell.style = "percent"
                        elif col_format == "number":
                            cell.style = "number"
                        else:
                            # Alternating row colors
                            if row_idx % 2 == 1:
                                cell.style = "alt_row"
                            else:
                                cell.style = "data"
                    else:
                        cell.style = "data"

                current_row += 1

        return ws

    def add_chart(
        self,
        sheet_name: str,
        chart_type: str = "bar",
        title: str = "",
        data_range: tuple = None,  # (min_col, min_row, max_col, max_row)
        categories_col: int = 1,
        position: str = "E2",
    ):
        """
        Add a chart to a worksheet.

        Args:
            sheet_name: Name of worksheet to add chart to
            chart_type: "bar" or "line"
            title: Chart title
            data_range: Tuple of (min_col, min_row, max_col, max_row)
            categories_col: Column number for category labels
            position: Cell position for chart (e.g., "E2")
        """
        if sheet_name not in self.wb.sheetnames:
            return

        ws = self.wb[sheet_name]

        if chart_type == "line":
            chart = LineChart()
        else:
            chart = BarChart()

        chart.title = title
        chart.style = 10

        if data_range:
            min_col, min_row, max_col, max_row = data_range
            data = Reference(ws, min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row)
            cats = Reference(ws, min_col=categories_col, min_row=min_row + 1, max_row=max_row)

            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)

        ws.add_chart(chart, position)

    def add_formula_row(
        self,
        sheet_name: str,
        row: int,
        formulas: Dict[int, str],  # col_idx -> formula
    ):
        """
        Add a row with formulas.

        Args:
            sheet_name: Worksheet name
            row: Row number
            formulas: Dict mapping column index to formula
        """
        if sheet_name not in self.wb.sheetnames:
            return

        ws = self.wb[sheet_name]

        for col_idx, formula in formulas.items():
            cell = ws.cell(row=row, column=col_idx)
            cell.value = formula
            cell.style = "highlight"

    def build_from_yaml(self, yaml_path: Path) -> "XLSXBuilder":
        """
        Build spreadsheet from YAML definition.

        YAML format:
        ```yaml
        title: "Spreadsheet Title"

        sheets:
          - name: "Data"
            title: "Sales Data"
            columns:
              - name: "Product"
                width: 20
              - name: "Revenue"
                width: 15
                format: currency
              - name: "Growth"
                width: 12
                format: percent
            rows:
              - ["Product A", 10000, 0.15]
              - ["Product B", 15000, 0.22]
            chart:
              type: bar
              title: "Revenue by Product"
              position: "E2"
        ```
        """
        import yaml

        with open(yaml_path, "r") as f:
            data = yaml.safe_load(f)

        for sheet_data in data.get("sheets", []):
            # Parse columns
            columns = []
            for col_def in sheet_data.get("columns", []):
                if isinstance(col_def, str):
                    columns.append(Column(name=col_def))
                else:
                    columns.append(
                        Column(
                            name=col_def.get("name", ""),
                            width=col_def.get("width", 12),
                            format=col_def.get("format", "general"),
                        )
                    )

            # Add sheet
            self.add_sheet(
                name=sheet_data.get("name", "Sheet"),
                title=sheet_data.get("title", ""),
                columns=columns,
                rows=sheet_data.get("rows", []),
            )

            # Add chart if defined
            chart_def = sheet_data.get("chart")
            if chart_def:
                # Calculate data range from rows
                num_rows = len(sheet_data.get("rows", []))
                num_cols = len(columns)
                header_row = 3 if sheet_data.get("title") else 1

                self.add_chart(
                    sheet_name=sheet_data.get("name", "Sheet"),
                    chart_type=chart_def.get("type", "bar"),
                    title=chart_def.get("title", ""),
                    data_range=(2, header_row, num_cols, header_row + num_rows),
                    position=chart_def.get("position", "E2"),
                )

        return self

    def build_from_dict(self, data: Dict[str, Any]) -> "XLSXBuilder":
        """
        Build spreadsheet from dictionary data.

        Simple format for quick spreadsheets:
        ```python
        {
            "Sheet1": {
                "headers": ["Name", "Value"],
                "rows": [
                    ["Item A", 100],
                    ["Item B", 200],
                ]
            }
        }
        ```
        """
        for sheet_name, sheet_data in data.items():
            headers = sheet_data.get("headers", [])
            rows = sheet_data.get("rows", [])

            columns = [Column(name=h, width=15) for h in headers]

            self.add_sheet(
                name=sheet_name,
                columns=columns,
                rows=rows,
            )

        return self

    def save(self, output_path: Path) -> Path:
        """Save spreadsheet to file."""
        self._remove_default_sheet()
        output_path.parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(str(output_path))
        return output_path


def build_xlsx(
    source_path: Path,
    output_path: Path,
    theme: "Theme" = None,
) -> Path:
    """
    Build XLSX from YAML source.

    Args:
        source_path: Path to .yaml file
        output_path: Where to save .xlsx
        theme: Optional theme for styling

    Returns:
        Path to generated XLSX
    """
    builder = XLSXBuilder(theme=theme)
    builder.build_from_yaml(source_path)
    return builder.save(output_path)
